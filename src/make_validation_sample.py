"""Draw the stratified sample for Day 3 hand-validation.

Reads the full tag table plus each charity's register text and writes two
files to outputs/validation/:

  labelling_sheet.xlsx  - 150 charities for hand-labelling, with dropdown
                          menus for the labels. The model's answers are
                          deliberately absent so the labelling is blind.
  sample_key.csv        - the model's answers for the same 150 rows, read
                          only by score_validation.py after labelling.

The sampling design is fixed here, before any labelling happens, so the
protocol can't quietly bend to fit the numbers later:

- n = 150 with a fixed random seed (recorded in the key file).
- Strata: overseas_engagement (3) x sdg_confidence (3) = 9 cells, allocated
  proportionally to the population with a floor of 6 rows per cell, so the
  rare low-confidence cells still get enough rows to say something about.
- Within each cell, rows are spread across primary_sdg in proportion to the
  cell's own mix (random within each goal), so the sample mirrors the
  population instead of drifting toward whatever the random draw hits.
- Finally, swaps within a cell guarantee all 17 goals appear at least once
  (goals like SDG 7, with only 35 charities, would otherwise never show up).

Each row also gets a population weight (its cell's population divided by
its cell's sample count), so the scorer can undo the floor's oversampling
of small cells and report a population-level accuracy estimate alongside
the raw sample accuracy.

Run from the repo root:  .venv/bin/python src/make_validation_sample.py
"""

import argparse
import random
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from classify_prompt import _SDG_REFERENCE, OBJECTS_CHAR_CAP, SDG_TITLES

REPO_ROOT = Path(__file__).resolve().parent.parent
TAGS_PATH = REPO_ROOT / "data" / "processed" / "sdg_tags.csv"
TEXT_PATH = REPO_ROOT / "data" / "processed" / "international.csv"
OUT_DIR = REPO_ROOT / "outputs" / "validation"
SHEET_PATH = OUT_DIR / "labelling_sheet.xlsx"
KEY_PATH = OUT_DIR / "sample_key.csv"

SAMPLE_SIZE = 150
SEED = 20260710  # fixed so the exact same sample can be re-drawn
CELL_FLOOR = 6   # minimum rows per engagement x confidence cell

ENGAGEMENT_CLASSES = [
    "operates_directly_abroad",
    "funds_partners_abroad",
    "uk_fundraising_only",
]

# The engagement definitions the model was given, restated for the labeller
# so both judge by the same rulebook (kept in step with classify_prompt.py).
LABELLING_GUIDE = [
    "HOW TO LABEL (read this once before starting)",
    "",
    "Judge each charity ONLY from the text in its row - the same text the "
    "model saw. Do not look anything up, and ignore anything you happen to "
    "know about a named charity.",
    "",
    "my_primary_sdg: the single UN Sustainable Development Goal (1-17) that "
    "best fits the charity's work. Use the reference below.",
    "",
    "my_alt_sdg: OPTIONAL - only for genuinely dual-purpose charities where "
    "a second goal would be equally correct as primary. Leave blank "
    "otherwise. This is not 'second most relevant'; it means 'either of "
    "these two would be a correct primary'.",
    "",
    "my_engagement: choose exactly one.",
    "- operates_directly_abroad: the charity itself runs activities or has "
    "staff/projects in other countries.",
    "- funds_partners_abroad: the charity mainly gives grants to, or works "
    "through, partner organisations overseas. A 'Friends of X' charity "
    "raising money for a named institution abroad belongs here.",
    "- uk_fundraising_only: the text describes UK fundraising or "
    "awareness-raising with no indication of how (or whether) money or "
    "activity reaches other countries. Default sparse cases here ONLY if "
    "there is no overseas signal at all.",
    "",
    "my_notes: one line on every call you found ambiguous, so the process "
    "is auditable. Leave blank when the call was easy.",
    "",
    "Tips: long text is easier to read in the formula bar (click the cell), "
    "or widen the row. Save the file as you go.",
]


def load_population() -> pd.DataFrame:
    tags = pd.read_csv(TAGS_PATH)
    text = pd.read_csv(TEXT_PATH)
    cols = ["organisation_number", "charity_name", "charity_activities",
            "charitable_objects"]
    df = tags.merge(text[cols], on="organisation_number", how="left")
    if len(df) != len(tags):
        sys.exit("join changed the row count - investigate before sampling")
    return df


def allocate(cell_sizes: pd.Series) -> dict:
    """Split SAMPLE_SIZE across cells: proportional, with a floor per cell."""
    total = cell_sizes.sum()
    alloc = {cell: max(CELL_FLOOR, round(SAMPLE_SIZE * n / total))
             for cell, n in cell_sizes.items()}
    # Rounding and the floor rarely land exactly on 150, so nudge the
    # largest cells (they can spare rows) until the total is right.
    while sum(alloc.values()) > SAMPLE_SIZE:
        biggest = max(alloc, key=alloc.get)
        alloc[biggest] -= 1
    while sum(alloc.values()) < SAMPLE_SIZE:
        biggest = max(cell_sizes.index, key=lambda c: cell_sizes[c] - alloc[c])
        alloc[biggest] += 1
    return alloc


def pick_cell_rows(cell_df: pd.DataFrame, quota: int, rng: random.Random) -> list:
    """Pick `quota` org numbers from one cell, proportionally across SDGs.

    Each goal gets seats in proportion to its share of the cell (largest
    remainder for the leftovers), then rows are drawn at random within each
    goal. This keeps the sample's SDG mix close to the population's, which
    is what lets a simple per-cell weight recover a population estimate.
    """
    sizes = cell_df["primary_sdg"].value_counts()
    exact = sizes * quota / sizes.sum()
    seats = exact.astype(int)
    # hand the leftover seats to the largest fractional remainders
    for sdg in (exact - seats).sort_values(ascending=False).index:
        if seats.sum() >= quota:
            break
        seats[sdg] += 1
    picked = []
    for sdg, k in seats.items():
        ids = list(cell_df.loc[cell_df["primary_sdg"] == sdg,
                               "organisation_number"])
        rng.shuffle(ids)
        picked += ids[:k]
    return picked


def draw_sample(df: pd.DataFrame, seed: int, verbose: bool = True) -> pd.DataFrame:
    """The full draw: allocate, pick per cell, then guarantee SDG coverage."""
    rng = random.Random(seed)
    df = df.copy()
    df["stratum"] = df["overseas_engagement"] + "|" + df["sdg_confidence"]
    cell_sizes = df["stratum"].value_counts()
    alloc = allocate(cell_sizes)

    picked_ids = []
    for cell in sorted(alloc):  # sorted so the draw order is reproducible
        cell_df = df[df["stratum"] == cell]
        picked_ids += pick_cell_rows(cell_df, alloc[cell], rng)
    sample = df[df["organisation_number"].isin(picked_ids)].copy()

    # Coverage pass: if a goal is missing entirely, swap it in for a row of
    # an over-represented goal from the same cell (cell counts stay intact).
    for sdg in range(1, 18):
        if (sample["primary_sdg"] == sdg).any():
            continue
        candidates = df[(df["primary_sdg"] == sdg)
                        & ~df["organisation_number"].isin(picked_ids)]
        if candidates.empty:
            print(f"note: no population rows at all for SDG {sdg}")
            continue
        incoming = candidates.sample(1, random_state=rng.randrange(2**32)).iloc[0]
        in_cell = sample[sample["stratum"] == incoming["stratum"]]
        common = in_cell["primary_sdg"].value_counts().idxmax()
        outgoing = in_cell[in_cell["primary_sdg"] == common].sample(
            1, random_state=rng.randrange(2**32)).iloc[0]
        sample = sample[sample["organisation_number"]
                        != outgoing["organisation_number"]]
        sample = pd.concat([sample, incoming.to_frame().T])
        picked_ids.append(incoming["organisation_number"])
        if verbose:
            print(f"coverage swap in {incoming['stratum']}: SDG {sdg} in, "
                  f"one SDG {common} row out")

    # Population weight: how many population charities each sampled row
    # stands for (its cell's population / its cell's sample count). The
    # scorer uses these to undo the floor's oversampling of small cells.
    samp_cells = sample["stratum"].value_counts()
    sample["weight"] = [cell_sizes[s] / samp_cells[s]
                        for s in sample["stratum"]]

    # Shuffle the final row order so strata can't be guessed while labelling.
    sample = sample.sample(frac=1, random_state=seed).reset_index(drop=True)
    return sample


def write_sheet(sample: pd.DataFrame) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "labelling"
    header = ["organisation_number", "charity_name", "charity_activities",
              "charitable_objects", "my_primary_sdg", "my_alt_sdg",
              "my_engagement", "my_notes"]
    ws.append(header)

    def clean(value) -> str:
        # pandas gives NaN (a float) for empty CSV cells
        return value.strip() if isinstance(value, str) else ""

    for _, row in sample.iterrows():
        ws.append([int(row["organisation_number"]),
                   clean(row["charity_name"]),
                   clean(row["charity_activities"]),
                   clean(row["charitable_objects"])[:OBJECTS_CHAR_CAP],
                   None, None, None, None])

    base = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    widths = {"A": 12, "B": 30, "C": 55, "D": 70, "E": 14, "F": 12,
              "G": 26, "H": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    n_rows = len(sample) + 1
    for r in range(2, n_rows + 1):
        ws.row_dimensions[r].height = 90
        for cell in ws[r]:
            cell.font = base
            cell.alignment = wrap
    ws.freeze_panes = "B2"

    # Dropdowns: SDG numbers for the two goal columns, the three class
    # names for the engagement column. These prevent typos the scorer
    # would otherwise have to guess about.
    dv_sdg = DataValidation(
        type="list",
        formula1='"' + ",".join(str(n) for n in range(1, 18)) + '"',
        allow_blank=True, showDropDown=False)
    dv_eng = DataValidation(
        type="list",
        formula1='"' + ",".join(ENGAGEMENT_CLASSES) + '"',
        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_sdg)
    ws.add_data_validation(dv_eng)
    dv_sdg.add(f"E2:F{n_rows}")
    dv_eng.add(f"G2:G{n_rows}")

    # Reference tab: labelling instructions plus the same SDG reference the
    # model worked from, so labeller and model judge by one rulebook.
    ref = wb.create_sheet("reference")
    ref.column_dimensions["A"].width = 100
    lines = LABELLING_GUIDE + ["", "-" * 60, ""]
    lines += [f"{n}. {title}" for n, title in SDG_TITLES.items()]
    lines += ["", "-" * 60, ""] + _SDG_REFERENCE.splitlines()
    for i, line in enumerate(lines, start=1):
        cell = ref.cell(row=i, column=1, value=line)
        cell.font = header_font if i == 1 else base
        cell.alignment = wrap

    wb.save(SHEET_PATH)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true",
                        help="overwrite an existing labelling sheet")
    args = parser.parse_args()
    if SHEET_PATH.exists() and not args.force:
        sys.exit(f"{SHEET_PATH} already exists - if labelling has started, "
                 "overwriting would destroy work. Pass --force to redraw.")

    df = load_population()
    print(f"population: {len(df):,} tagged charities")

    sample = draw_sample(df, SEED)
    redraw = draw_sample(df, SEED, verbose=False)
    assert list(sample["organisation_number"]) == \
        list(redraw["organisation_number"]), "draw is not deterministic!"
    assert len(sample) == SAMPLE_SIZE

    # --- printed sanity checks ------------------------------------------
    print(f"\nsample: {len(sample)} rows, seed {SEED} (re-draw identical)")
    print("\nrows per stratum (population share -> sample share):")
    pop_share = (df["overseas_engagement"] + "|" + df["sdg_confidence"]
                 ).value_counts(normalize=True)
    for cell, n in sample["stratum"].value_counts().sort_index().items():
        print(f"  {cell:<40} {pop_share[cell]:>5.1%} -> {n / len(sample):.1%}"
              f"  ({n} rows)")
    covered = sorted(int(s) for s in sample["primary_sdg"].unique())
    print(f"\nSDGs covered: {len(covered)}/17 {covered}")
    counts = sample["primary_sdg"].astype(int).value_counts().sort_index()
    print("sample SDG counts:", {int(k): int(v) for k, v in counts.items()})
    print("weights: min {:.1f}, max {:.1f}, sum {:.0f} (~population size)"
          .format(sample["weight"].min(), sample["weight"].max(),
                  sample["weight"].sum()))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    key_cols = sample[[
        "organisation_number", "stratum", "weight", "primary_sdg",
        "secondary_sdgs", "sdg_confidence", "overseas_engagement",
        "engagement_confidence",
    ]].rename(columns={
        "primary_sdg": "model_primary_sdg",
        "secondary_sdgs": "model_secondary_sdgs",
        "sdg_confidence": "model_sdg_confidence",
        "overseas_engagement": "model_overseas_engagement",
        "engagement_confidence": "model_engagement_confidence",
    })
    key_cols.insert(1, "seed", SEED)
    key_cols.to_csv(KEY_PATH, index=False)
    print(f"\nSaved key to {KEY_PATH}")

    write_sheet(sample)
    print(f"Saved labelling sheet to {SHEET_PATH}")
    print("\nNext: fill in my_primary_sdg, my_engagement (and my_alt_sdg / "
          "my_notes where needed)\nfor every row, save, then run "
          ".venv/bin/python src/score_validation.py")


if __name__ == "__main__":
    main()
