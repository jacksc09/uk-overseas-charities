"""A small local web app for hand-labelling the validation sample.

Reads outputs/validation/labelling_sheet.xlsx (never the answer key) and
shows one charity at a time with keyboard-first controls, which is much
faster than editing spreadsheet cells. Every label is written straight
back into the same xlsx the scorer reads, so Excel and this app are
interchangeable - you can switch between them at any point (just not at
the same time: close the file in Excel while this app is running).

Safety nets, because 150 hand labels are hours of work:
- on startup a one-off backup copy of the sheet is written next to it;
- every save is also appended to a plain-text journal file, so even a
  corrupted workbook could be replayed.

Run from the repo root:  .venv/bin/python src/label_ui.py
then open http://localhost:8765 (it opens itself by default).
"""

import argparse
import json
import shutil
import sys
import threading
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import openpyxl

from classify_prompt import _SDG_REFERENCE, SDG_TITLES

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SHEET = REPO_ROOT / "outputs" / "validation" / "labelling_sheet.xlsx"

ENGAGEMENT_CLASSES = [
    "operates_directly_abroad",
    "funds_partners_abroad",
    "uk_fundraising_only",
]

# Sheet columns, by header name (resolved to letters at startup).
LABEL_COLS = ["my_primary_sdg", "my_alt_sdg", "my_engagement", "my_notes"]


class SheetStore:
    """Owns the workbook: reads all rows, writes one row's labels at a time."""

    def __init__(self, path: Path):
        self.path = path
        self.lock = threading.Lock()  # one write at a time
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb["labelling"]
        headers = [c.value for c in self.ws[1]]
        self.col = {name: headers.index(name) + 1 for name in
                    ["organisation_number", "charity_name",
                     "charity_activities", "charitable_objects"] + LABEL_COLS}
        # organisation_number -> worksheet row index, for O(1) writes
        self.row_of = {self.ws.cell(r, self.col["organisation_number"]).value: r
                       for r in range(2, self.ws.max_row + 1)}
        self.journal = path.with_suffix(".journal.jsonl")

    def rows(self) -> list:
        out = []
        for r in range(2, self.ws.max_row + 1):
            get = lambda name: self.ws.cell(r, self.col[name]).value
            out.append({
                "orgno": get("organisation_number"),
                "name": get("charity_name") or "",
                "activities": get("charity_activities") or "",
                "objects": get("charitable_objects") or "",
                "primary": get("my_primary_sdg"),
                "alt": get("my_alt_sdg"),
                "engagement": get("my_engagement"),
                "notes": get("my_notes") or "",
            })
        return out

    def save_label(self, data: dict) -> None:
        """Validate one row's labels and write them through to disk."""
        orgno = data.get("orgno")
        if orgno not in self.row_of:
            raise ValueError(f"unknown organisation_number {orgno}")
        primary, alt = data.get("primary"), data.get("alt")
        engagement, notes = data.get("engagement"), data.get("notes") or ""
        for value, label in ((primary, "primary"), (alt, "alt")):
            if value is not None and value not in range(1, 18):
                raise ValueError(f"{label} SDG must be 1-17, got {value!r}")
        if alt is not None and alt == primary:
            raise ValueError("alt SDG is only for a DIFFERENT, equally "
                             "correct primary - it can't equal the primary")
        if engagement is not None and engagement not in ENGAGEMENT_CLASSES:
            raise ValueError(f"bad engagement value {engagement!r}")

        with self.lock:
            r = self.row_of[orgno]
            self.ws.cell(r, self.col["my_primary_sdg"], primary)
            self.ws.cell(r, self.col["my_alt_sdg"], alt)
            self.ws.cell(r, self.col["my_engagement"], engagement)
            self.ws.cell(r, self.col["my_notes"], notes or None)
            self.wb.save(self.path)  # raises if e.g. Excel has it locked
            with open(self.journal, "a") as f:
                f.write(json.dumps({"at": datetime.now().isoformat(),
                                    **data}) + "\n")

    def done_count(self) -> int:
        """A row is done when it has both a primary SDG and an engagement."""
        return sum(1 for r in range(2, self.ws.max_row + 1)
                   if self.ws.cell(r, self.col["my_primary_sdg"]).value
                   and self.ws.cell(r, self.col["my_engagement"]).value)


# ---------------------------------------------------------------------------
# The page itself. Everything is inline so this file is the whole app.
# ---------------------------------------------------------------------------

PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Hand-labelling</title>
<style>
  :root { --ink:#1a1a1a; --mut:#666; --line:#ddd; --accent:#2456a6;
          --good:#1d7a3d; --warn:#a63324; --bg:#f7f7f5; --card:#fff; }
  * { box-sizing:border-box; }
  body { margin:0; font:15px/1.5 -apple-system, "Segoe UI", sans-serif;
         color:var(--ink); background:var(--bg); }
  .wrap { display:grid; grid-template-columns: 1fr 340px; gap:16px;
          max-width:1280px; margin:0 auto; padding:16px; }
  header { grid-column:1/-1; display:flex; align-items:baseline; gap:16px; }
  header h1 { font-size:18px; margin:0; }
  #progress { color:var(--mut); }
  #saved { margin-left:auto; font-size:13px; color:var(--good); }
  #saved.err { color:var(--warn); font-weight:600; }
  .card { background:var(--card); border:1px solid var(--line);
          border-radius:10px; padding:18px 20px; }
  #charity h2 { margin:0 0 2px; font-size:20px; }
  #charity .orgno { color:var(--mut); font-size:13px; margin-bottom:12px; }
  #charity h3 { font-size:12px; text-transform:uppercase; letter-spacing:.06em;
                color:var(--mut); margin:14px 0 4px; }
  #charity .text { white-space:pre-wrap; max-height:180px; overflow-y:auto; }
  .controls { margin-top:16px; }
  .row-label { font-size:12px; text-transform:uppercase; letter-spacing:.06em;
               color:var(--mut); margin:12px 0 6px; }
  .sdg-grid { display:grid; grid-template-columns:repeat(9, 1fr); gap:6px; }
  .sdg-grid button { padding:7px 2px; border:1px solid var(--line);
        border-radius:7px; background:#fff; cursor:pointer; font-size:13px; }
  .sdg-grid button.sel-primary { background:var(--accent); color:#fff;
        border-color:var(--accent); font-weight:700; }
  .sdg-grid button.sel-alt { background:#dbe6f6; border-color:var(--accent); }
  .eng-row { display:flex; gap:8px; }
  .eng-row button { flex:1; padding:10px 6px; border:1px solid var(--line);
        border-radius:7px; background:#fff; cursor:pointer; font-size:13px; }
  .eng-row button.sel { background:var(--accent); color:#fff;
        border-color:var(--accent); font-weight:600; }
  .eng-row b { display:block; font-size:15px; }
  textarea { width:100%; min-height:52px; border:1px solid var(--line);
        border-radius:7px; padding:8px; font:inherit; resize:vertical; }
  .nav { display:flex; gap:8px; margin-top:14px; align-items:center; }
  .nav button { padding:9px 16px; border:1px solid var(--line);
        border-radius:7px; background:#fff; cursor:pointer; }
  .nav .next { background:var(--good); color:#fff; border-color:var(--good);
        font-weight:600; }
  .nav .hint { color:var(--mut); font-size:13px; margin-left:auto;
        text-align:right; }
  kbd { background:#eee; border:1px solid #ccc; border-radius:4px;
        padding:0 5px; font-size:12px; }
  #buffer { font-size:13px; color:var(--mut); min-height:1.2em; margin-top:6px; }
  #buffer b { color:var(--accent); font-size:15px; }
  aside .card { position:sticky; top:16px; max-height:calc(100vh - 32px);
        overflow-y:auto; font-size:13px; }
  aside h3 { margin:4px 0 8px; font-size:14px; }
  aside pre { white-space:pre-wrap; font:12px/1.45 inherit; color:#333; }
  #jump { width:100%; margin-top:8px; }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>Hand-labelling</h1>
    <span id="progress"></span>
    <span id="saved"></span>
  </header>

  <main>
    <div class="card" id="charity">
      <h2 id="c-name"></h2>
      <div class="orgno" id="c-orgno"></div>
      <h3>Activities</h3><div class="text" id="c-act"></div>
      <h3>Objects</h3><div class="text" id="c-obj"></div>
    </div>

    <div class="card controls">
      <div class="row-label">Primary SDG — type the number then
        <kbd>Enter</kbd>, or click. (<kbd>a</kbd> + number = alt SDG for
        genuinely dual-purpose charities only; <kbd>x</kbd> clears alt)</div>
      <div class="sdg-grid" id="sdg-grid"></div>
      <div id="buffer"></div>

      <div class="row-label">Overseas engagement</div>
      <div class="eng-row" id="eng-row">
        <button data-eng="operates_directly_abroad"><b>D</b>irect: runs its
          own activities/staff abroad</button>
        <button data-eng="funds_partners_abroad"><b>P</b>artners: grants to /
          works through orgs abroad</button>
        <button data-eng="uk_fundraising_only"><b>U</b>K only: no overseas
          mechanism in the text</button>
      </div>

      <div class="row-label">Notes — one line for every ambiguous call
        (<kbd>n</kbd> to focus, <kbd>Esc</kbd> to leave)</div>
      <textarea id="notes" placeholder="why this call was hard (leave blank if easy)"></textarea>

      <div class="nav">
        <button id="prev">← prev</button>
        <button id="next" class="next">next unlabelled →</button>
        <span class="hint"><kbd>←</kbd>/<kbd>→</kbd> move ·
          <kbd>Enter</kbd> (empty buffer) = next</span>
      </div>
    </div>
  </main>

  <aside>
    <div class="card">
      <h3>Reference — one rulebook for you and the model</h3>
      <pre id="reference"></pre>
      <select id="jump"></select>
    </div>
  </aside>
</div>

<script>
"use strict";
const TITLES = __TITLES__;
let rows = [], cur = 0, buffer = "", altMode = false, noteTimer = null;

const $ = id => document.getElementById(id);

async function boot() {
  const resp = await fetch("/api/rows");
  const data = await resp.json();
  rows = data.rows;
  $("reference").textContent = data.reference;
  const grid = $("sdg-grid");
  for (let n = 1; n <= 17; n++) {
    const b = document.createElement("button");
    b.textContent = n;
    b.title = n + ". " + TITLES[n];
    b.onclick = () => { setSDG(n, altMode); altMode = false; drawBuffer(); };
    grid.appendChild(b);
  }
  const jump = $("jump");
  rows.forEach((r, i) => {
    const o = document.createElement("option");
    o.value = i; o.textContent = (i + 1) + ". " + r.name.slice(0, 40);
    jump.appendChild(o);
  });
  jump.onchange = () => show(+jump.value);
  // resume where labelling left off
  const first = rows.findIndex(r => !(r.primary && r.engagement));
  show(first === -1 ? 0 : first);
}

function show(i) {
  cur = Math.max(0, Math.min(rows.length - 1, i));
  const r = rows[cur];
  $("c-name").textContent = (cur + 1) + ". " + r.name;
  $("c-orgno").textContent = "organisation_number " + r.orgno;
  $("c-act").textContent = r.activities || "(none provided)";
  $("c-obj").textContent = r.objects || "(none provided)";
  $("notes").value = r.notes || "";
  $("jump").value = cur;
  buffer = ""; altMode = false;
  drawButtons(); drawBuffer(); drawProgress();
}

function drawButtons() {
  const r = rows[cur];
  document.querySelectorAll("#sdg-grid button").forEach((b, idx) => {
    b.className = "";
    if (r.primary === idx + 1) b.className = "sel-primary";
    else if (r.alt === idx + 1) b.className = "sel-alt";
  });
  document.querySelectorAll("#eng-row button").forEach(b => {
    b.classList.toggle("sel", b.dataset.eng === r.engagement);
  });
}

function drawProgress() {
  const done = rows.filter(r => r.primary && r.engagement).length;
  $("progress").textContent =
    done + "/" + rows.length + " labelled — row " + (cur + 1);
}

function drawBuffer() {
  $("buffer").innerHTML = altMode
    ? "alt SDG: <b>" + (buffer || "…") + "</b> (Enter to set)"
    : buffer ? "primary SDG: <b>" + buffer + "</b> (Enter to set)" : "";
}

function setSDG(n, asAlt) {
  const r = rows[cur];
  if (asAlt) {
    if (n === r.primary) { flash("alt must differ from primary", true); return; }
    r.alt = n;
  } else {
    r.primary = n;
    if (r.alt === n) r.alt = null;
  }
  drawButtons(); drawProgress(); save();
}

function setEng(cls) {
  rows[cur].engagement = cls;
  drawButtons(); drawProgress(); save();
}

async function save() {
  const r = rows[cur];
  try {
    const resp = await fetch("/api/label", {
      method: "POST", headers: {"Content-Type": "application/json"},
      body: JSON.stringify({orgno: r.orgno, primary: r.primary, alt: r.alt,
                            engagement: r.engagement, notes: r.notes})
    });
    const out = await resp.json();
    if (!resp.ok) throw new Error(out.error);
    flash("saved ✓");
  } catch (e) {
    flash("NOT SAVED: " + e.message + " (is the file open in Excel?)", true);
  }
}

function flash(msg, isErr) {
  const el = $("saved");
  el.textContent = msg;
  el.className = isErr ? "err" : "";
  if (!isErr) setTimeout(() => { if (el.textContent === msg) el.textContent = ""; }, 1500);
}

function nextUnlabelled() {
  const after = rows.findIndex((r, i) => i > cur && !(r.primary && r.engagement));
  if (after !== -1) return show(after);
  const any = rows.findIndex(r => !(r.primary && r.engagement));
  if (any !== -1) return show(any);
  show(Math.min(cur + 1, rows.length - 1));
  flash("all 150 rows are labelled 🎉");
}

$("prev").onclick = () => show(cur - 1);
$("next").onclick = nextUnlabelled;
document.querySelectorAll("#eng-row button").forEach(b => {
  b.onclick = () => setEng(b.dataset.eng);
});
$("notes").addEventListener("input", () => {
  rows[cur].notes = $("notes").value;
  clearTimeout(noteTimer);
  noteTimer = setTimeout(save, 600);   // save notes after a typing pause
});
$("notes").addEventListener("keydown", e => {
  if (e.key === "Escape") $("notes").blur();
  e.stopPropagation();                 // typing notes must not trigger hotkeys
});

document.addEventListener("keydown", e => {
  if (e.target.tagName === "TEXTAREA" || e.target.tagName === "SELECT") return;
  if (e.key >= "0" && e.key <= "9") {
    buffer = (buffer + e.key).slice(-2);
    if (+buffer < 1 || +buffer > 17) buffer = e.key;   // e.g. "18" -> "8"
    drawBuffer();
  } else if (e.key === "Enter") {
    if (buffer) { setSDG(+buffer, altMode); buffer = ""; altMode = false; drawBuffer(); }
    else nextUnlabelled();
  } else if (e.key === "Backspace") {
    buffer = buffer.slice(0, -1); drawBuffer();
  } else if (e.key === "a" || e.key === "A") {
    altMode = !altMode; drawBuffer();
  } else if (e.key === "x" || e.key === "X") {
    rows[cur].alt = null; drawButtons(); save();
  } else if (e.key === "d" || e.key === "D") setEng("operates_directly_abroad");
  else if (e.key === "p" || e.key === "P") setEng("funds_partners_abroad");
  else if (e.key === "u" || e.key === "U") setEng("uk_fundraising_only");
  else if (e.key === "n" || e.key === "N") { e.preventDefault(); $("notes").focus(); }
  else if (e.key === "ArrowLeft") show(cur - 1);
  else if (e.key === "ArrowRight") show(cur + 1);
});

boot();
</script>
</body>
</html>
"""


class Handler(BaseHTTPRequestHandler):
    store: SheetStore = None  # set in main()

    def _send(self, code: int, body: bytes, ctype: str) -> None:
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _json(self, code: int, obj: dict) -> None:
        self._send(code, json.dumps(obj).encode(), "application/json")

    def do_GET(self):
        if self.path == "/":
            page = PAGE.replace("__TITLES__", json.dumps(SDG_TITLES))
            self._send(200, page.encode(), "text/html; charset=utf-8")
        elif self.path == "/api/rows":
            self._json(200, {"rows": self.store.rows(),
                             "reference": _SDG_REFERENCE})
        else:
            self._json(404, {"error": "not found"})

    def do_POST(self):
        if self.path != "/api/label":
            return self._json(404, {"error": "not found"})
        length = int(self.headers.get("Content-Length", 0))
        try:
            data = json.loads(self.rfile.read(length))
            self.store.save_label(data)
            self._json(200, {"ok": True, "done": self.store.done_count()})
        except Exception as exc:  # report the reason instead of a stack trace
            self._json(400, {"error": str(exc)})

    def log_message(self, *args):  # keep the terminal quiet while labelling
        pass


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sheet", type=Path, default=DEFAULT_SHEET)
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--no-open", action="store_true",
                        help="don't open the browser automatically")
    args = parser.parse_args()
    if not args.sheet.exists():
        sys.exit(f"{args.sheet} not found - run make_validation_sample.py first")

    backup = args.sheet.with_suffix(".startup-backup.xlsx")
    shutil.copy2(args.sheet, backup)
    print(f"backup of the sheet saved to {backup.name}")

    Handler.store = SheetStore(args.sheet)
    print(f"{Handler.store.done_count()}/{len(Handler.store.row_of)} rows "
          "already labelled")

    url = f"http://localhost:{args.port}"
    server = ThreadingHTTPServer(("localhost", args.port), Handler)
    print(f"labelling at {url}  (Ctrl+C to stop; progress saves as you go)")
    if not args.no_open:
        threading.Timer(0.5, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nstopped - your labels are saved in the sheet")


if __name__ == "__main__":
    main()
